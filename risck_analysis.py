==========================================================
IMPORTS
==========================================================
import json
import os
import time
from pathlib import Path

import pandas as pd
from openai import OpenAI
from difflib import SequenceMatcher
from difflib import get_close_matches
from openai import RateLimitError

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

==========================================================
CONSTANTS / CONFIGURATIONS
==========================================================
MODEL_RISCO = "gpt-4.1-mini"

CONTROLE_PADRAO = (
"Realizar o C.F.E. utilizar os EPI´s: "
"protetor auricular, óculos de segurança, "
"botinas e luvas Hyflex"
)

CLIENTE = OpenAI(api_key=r"sk-**************")

GRAVIDADES_VALIDAS = {
"Primeiros socorros",
"Incidente com médico",
"Incidente sem afastamento",
"Incidente com afastamento",
"Fatalidade",
}

PROBABILIDADES_VALIDAS = {
"Muito Improvável",
"Improvável",
"Possível",
"Provável",
"Muito Provável",
}

==========================================================
PROMPTS FOR OPEN AI
==========================================================
PROMPT_BASE = f"""
Você é um especialista em análise de riscos de manutenção industrial.

Analise a atividade abaixo e responda EXCLUSIVAMENTE no formato JSON.
Qualquer resposta fora do JSON será considerada inválida.

REGRAS IMPORTANTES:

"riscos" é um TEXTO multilinha (uma linha por risco, sem numeração)

"controle" é um TEXTO ÚNICO

"gravidade" e "probabilidade" DEVEM ser escolhidas EXATAMENTE das opções listadas

NÃO invente categorias

NÃO reescreva nomes

NÃO justifique fora dos campos

A gravidade e a probabilidade DEVEM ser proporcionais
APENAS aos riscos físicos imediatos listados,
considerando máquina parada, bloqueada e EPI padrão.

OPÇÕES DE GRAVIDADE (use exatamente uma):

Primeiros socorros

Incidente com médico

Incidente sem afastamento

Incidente com afastamento

Fatalidade

OPÇÕES DE PROBABILIDADE (use exatamente uma):

Muito Improvável

Improvável

Possível

Provável

Muito Provável

FORMATO DE SAÍDA (JSON PURO):
{{
"riscos": "texto",
"gravidade": "uma das opções",
"probabilidade": "uma das opções",
"controle": "texto"
}}

CRITÉRIO DE EXISTÊNCIA DE RISCO:
Um risco SÓ EXISTE se houver exposição física direta e imediata do operador a:

Energia elétrica não isolada

Partes móveis em movimento

Massa suspensa ou possibilidade real de queda

Pontos de esmagamento ou cisalhamento acessíveis

Superfícies cortantes, quentes, sob pressão ou agentes agressivos

CRITÉRIOS OBJETIVOS ADICIONAIS (OBRIGATÓRIOS):

Cavacos, rebarbas ou limalha:
Sempre que houver limpeza, remoção ou contato manual,
considerar OBRIGATORIAMENTE:
Corte ou perfuração nas mãos

1.1) Óleo ou fluido:
Considerar contato com resíduo.
Considerar escorregamento SOMENTE se houver alta probabilidade de derramamento.

Montagem ou desmontagem:
Considerar choque mecânico, impacto ou prensagem leve.

Escorregamento:
Considerar SOMENTE se houver alta probabilidade de superfície escorregadia.
Não considerar em atividades secas ou externas.

Atividades externas isoladas:
Se não houver contato com partes móveis, resíduos ou energia,
classificar obrigatoriamente como:
Sem riscos relevantes

CONTEXTO FIXO DO PROCEDIMENTO:

C.F.E. realizado

EPI padrão em uso

Máquina parada e bloqueada quando necessário

FORMATO DO CAMPO "riscos":

Listar APENAS riscos físicos imediatos e reais

Sem explicações

Sem a palavra "risco"

Um risco por linha

Se não houver risco, retornar EXATAMENTE:
Sem riscos relevantes

REGRAS PARA O CONTROLE

CONTROLE PADRÃO (texto fixo):
Realizar o C.F.E. utilizar os EPI´s: protetor auricular, óculos de segurança, botinas e luvas Hyflex

LÓGICA DE DECISÃO:

Se "riscos" for "Sem riscos relevantes", usar EXATAMENTE o controle padrão

Se houver riscos, adicionar SOMENTE ações físicas ou EPIs estritamente necessários

Nunca criar mais de um controle

Priorizar o mínimo absoluto de controles

CRITÉRIOS OBJETIVOS PARA CONTROLE:

Montagem/desmontagem: verificar estado das ferramentas

Deslocamento/transporte: verificar e desobstruir o caminho

Movimento linear/força: travar movimento e manter mãos fora da zona de esmagamento

Óleo: em caso de derramamento, realizar limpeza e descarte correto

Queda de componentes: apoiar ou sustentar antes da liberação

Atividade:
{{ATIVIDADE}}

"""

==========================================================
HELPER FUNCTIONS (PURE)
==========================================================
def normalizar_texto(texto: str) -> str:
if not texto:
return ""

text
texto = texto.strip()
texto = texto.replace("–", "-").replace("—", "-")
texto = " ".join(texto.split())  # remove duplicate spaces

return texto
def validar_valor(
valor: str,
opcoes_validas: set[str],
campo: str,
cutoff: float = 0.7,
) -> str:
"""
Validates and tries to correct values outside the allowed list.
"""

text
valor = normalizar_texto(valor)

if valor in opcoes_validas:
    return valor

sugestoes = get_close_matches(
    valor,
    opcoes_validas,
    n=1,
    cutoff=cutoff
)
def matriz_risco(gravidade: str, probabilidade: str) -> str:
"""
Returns the risk status based on the matrix
"""

text
if gravidade == "Primeiros socorros":
    if probabilidade in {"Muito Improvável", "Improvável", "Possível"}:
        return "Trivial"
    return "Tolerável"

if gravidade == "Incidente com médico":
    if probabilidade == "Muito Improvável":
        return "Trivial"
    if probabilidade in {"Improvável", "Possível"}:
        return "Tolerável"
    return "Substancial"

if gravidade == "Incidente sem afastamento":
    if probabilidade == "Muito Improvável":
        return "Trivial"
    if probabilidade == "Improvável":
        return "Tolerável"
    if probabilidade in {"Possível", "Provável"}:
        return "Substancial"
    return "Intolerável"

if gravidade == "Incidente com afastamento":
    if probabilidade == "Muito Improvável":
        return "Tolerável"
    if probabilidade in {"Improvável", "Possível"}:
        return "Substancial"
    return "Intolerável"

if gravidade == "Fatalidade":
    if probabilidade == "Muito Improvável":
        return "Tolerável"
    if probabilidade == "Improvável":
        return "Substancial"
    return "Intolerável"

return "Indefinido"
def formatacoa_por_status(status):
"""
Returns fill and font colors according to risk status.
"""

text
cores = {
    "Trivial": {
        "fill": "00b050",   # Light green
        "font": "963634"    # Orange
    },
    "Tolerável": {
        "fill": "ffff00",   # Yellow
        "font": "ebebeb"    # GRAY
    },
    "Substancial": {
        "fill": "FFC000",   # Orange
        "font": "ebebeb"    # GRAY
    },
    "Intolerável": {
        "fill": "FF0000",   # Red
        "font": "FFFF00"    # Yellow
    }
}

return cores.get(
    status,
    {"fill": "FFFFFF", "font": "000000"}  # fallback
)
==========================================================
FUNCTIONS THAT USE API / I-O
==========================================================
#Takes a maintenance activity and a prompt (script for AI) and returns a value
def gerar_resposta(atividade, prompt):
response = CLIENTE.chat.completions.create(
model=MODEL_RISCO,
messages=[
{"role": "system", "content": prompt},
{"role": "user", "content": atividade},
],
temperature=0,
)

text
conteudo = response.choices[0].message.content
return json.loads(conteudo)   
==========================================================
MAIN FUNCTION
==========================================================
def main() -> None:

text
base_path = Path(r"PATH_TO_YOUR_FILE.xlsx")

df = pd.read_excel(base_path,
    sheet_name="FORMULÁRIO AR",
    engine="openpyxl"
)

resultados_por_atividade = []

#Remove header/legend lines and get activity column
atividades = df.iloc[5:-5, 6].dropna()
print(atividades)

   

for atividade_atual in atividades:
    print(f"Current procedure: {atividade_atual}")

    try:
        resposta = gerar_resposta(atividade_atual, PROMPT_BASE)
    except RateLimitError as e:
        print("Rate limit reached. Ending execution.")
        break

    resultados = {
        "riscos": resposta["riscos"],
        "gravidade": resposta["gravidade"],
        "probabilidade": resposta["probabilidade"],
        "controle": resposta["controle"],
    }

    # 🔒 Validations / normalizations
    resultados["gravidade"] = validar_valor(
        resultados["gravidade"],
        GRAVIDADES_VALIDAS,
        campo="Gravidade",
    )

    resultados["probabilidade"] = validar_valor(
        resultados["probabilidade"],
        PROBABILIDADES_VALIDAS,
        campo="Probabilidade",
    )

    resultados["status"] = matriz_risco(
        resultados["gravidade"],
        resultados["probabilidade"]
    )

    print(resultados)

    resultados_por_atividade.append({
        "atividade": atividade_atual,
        **resultados
    })

    time.sleep(20)        #limit 3 per minute



wb = load_workbook(base_path)
ws = wb["FORMULÁRIO AR"]

for i, item in enumerate(resultados_por_atividade):
    linha_excel = i + 7

    ws.cell(row=linha_excel, column=12).value = item["riscos"]          # L
    ws.cell(row=linha_excel, column=22).value = item["probabilidade"]  # V
    ws.cell(row=linha_excel, column=26).value = item["gravidade"]      # Z
    ws.cell(row=linha_excel, column=30).value = item["controle"]       # AD

    #Status with formatting
    celula_status = ws.cell(row=linha_excel, column=17)                #Q
    celula_status.value = resultados["status"]

    cores = formatacao_por_status(resultados["status"])

    if cores:
        # Cell background
        celula_status.fill = PatternFill(
            fill_type="solid",
            fgColor=cores["fill"]
        )

        # Font: bold, size 18, color according to status
        celula_status.font = Font(
            bold=True,
            size=18,
            color=cores["font"]
        )



wb.save(base_path)
print("Risk analysis successfully saved to file.")
==========================================================
ENTRY POINT
==========================================================
if name == "main":
main()
